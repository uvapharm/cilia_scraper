# Queries PubMed to find articles relating to primary. Returns in spreadsheet with
# LLm generated responses to predefined user questions

import requests
import glob
import sys
import bs4
import re
from time import sleep
import os
from datetime import datetime
from datetime import date
from openpyxl import load_workbook
from langchain.chat_models import ChatOpenAI
from joblib import load
import xlsxwriter
from langchain.schema import (
    HumanMessage,
    SystemMessage
)
from dotenv import load_dotenv

os.system('cls')
# NOTE: while the limit of records in 'fetch_pub_ids()' is 100,000, the limit for 'efetch' is "about 200", per the PubMED API site. Had to create several xml files and iterate through them.
pub_records_URL_start = 'http://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pubmed&retmode=xml&id='
output_dir = "" # choose output directory
results_dir = "" # choose results directory
results_files = "*.xlsx" # insert results directory ahead of .xlsx
xml_files = "*.xml" # insert xlm file directory ahead of .xml
xml_dir = "" # choose directory for the xml files


today_1 = date.today()
current_year = today_1.year
today_1 = today_1.strftime("%Y-%m-%d")
xlsx_file = results_dir + 'Publication_data_' + str(today_1) + '.xlsx'
file_name = 'Publication_data_' + str(today_1) + '.xlsx'

print('\n' +'    Processor Removing past xml files: ', end = '')
for m in os.listdir(xml_dir):
    os.remove(os.path.join(xml_dir, m))
sleep(2)
print('done')

# find the most recent results, for comparison to new data. If the most recent was run today, remove it.
if len(glob.glob(results_files)) != 0:
    most_recent_results = max(glob.glob(results_files), key=os.path.getctime)
    recent_results_split = os.path.split(most_recent_results)
    most_recent_results_file = recent_results_split[1]
    while True: # make sure the most recent  results file is not currently open.
        try:
            myfile = open(most_recent_results, 'r+')
            break
        except IOError:
            print('\n' + '    WARNING: ' + str(most_recent_results_file) + ' is open. Please close it and restart the program [F3 + ENTER]')
            quit()
    myfile.close()
    if most_recent_results_file == file_name:
        print('\n' + "    A file for today's data currently exists. Removing it now: ", end ='')
        os.remove(most_recent_results)
        print('that file has been removed.')

query_start = input('\n' + '    Enter the date from which you want the query to begin, in the YYYY/MM/DD format: ')
try:
    datetime.strptime(query_start, "%Y/%m/%d")
except ValueError:
    print('\n' + '        Incorrect date format, it should be YYYY/MM/DD. Restart [F3 - ENTER] and enter it in this format')
    quit()
pubs_query_url = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&term=%28%28%22Cilia%22[MeSH Terms]+OR+%22Ciliopathies%22[MeSH Terms]%29%29+AND+%28%28%22inhibit%22[Title/Abstract]+OR+inhibits[Title/Abstract]+OR+inhibited[Title/Abstract]+OR+inhibiting[Title/Abstract]+OR+%22promote%22[Title/Abstract]+OR+%22promotes%22[Title/Abstract]+OR+%22promoted%22[Title/Abstract]+OR+%22promoting%22[Title/Abstract]+OR+%22induce%22[Title/Abstract]+OR+%22induces%22[Title/Abstract]+OR+%22induced%22[Title/Abstract]+OR+%22inducing%22[Title/Abstract]+OR+%22regulates%22[Title/Abstract]+OR+%22alter%22[Title/Abstract]+OR+%22altering%22[Title/Abstract]%29%29+AND+%28%28%22small molecule inhibitors%22[Other Term]+OR+%22cilia length%22[Other Term]+OR+%22ciliogenesis%22[Other Term]+OR+%22primary cilia%22[Other Term]%29%29+AND+%28%28%22pharmacological%22[Title/Abstract]+OR+%22drug%22[Title/Abstract]+OR+inhibitor[Title/Abstract]%28%22Published Erratum%22[Publication Type]+OR+%22Retracted Publication%22[Publication Type]+OR+%22Retraction of Publication%22[Publication Type]+OR+Preprint[Publication Type]%29%29+AND+%28%28%22'+ query_start + '%22[PDAT]+:+%22' + str(current_year) + '/12/31%22[PDAT]%29%29&retmax=99999'
if len(glob.glob(results_files)) != 0:
    most_recent_results = max(glob.glob(results_files), key=os.path.getmtime)
    # create a list of PMIDs from the most recent search results
    global old_data
    old_data = []
    wb = load_workbook(most_recent_results)
    sheet = wb.active
    for row in sheet.iter_rows():
        old_data_row = []
        for cell in row:
            old_data_row += [cell.internal_value]
        old_data += [old_data_row]
    wb.close()
    old_pmids = []
    for x in range(1,len(old_data)):
        old_pmids.append(old_data[x][0])

start_time = datetime.now()
# load synonym dictionary
global dictionary
dictionary = []
wb = load_workbook(output_dir + 'Indications and synonyms.xlsx')
sheet = wb.active
for row in sheet.iter_rows():
    dictionary_row = []
    for cell in row:
        dictionary_row += [cell.internal_value]
    dictionary += [dictionary_row]
wb.close()

#load MeSH terms-to-R/T Type table
global type_dict
type_dict = []
wb = load_workbook(output_dir + '\Terms for R-T Type.xlsx')
sheet = wb.active
for row in sheet.iter_rows():
    type_dict_row = []
    for cell in row:
        type_dict_row += [cell.internal_value]
    type_dict += [type_dict_row]
wb.close()

def fetch_pub_ids():
    print('\n' + '    1. Getting pub IDs from PubMed: ', end = '')
    data_request = requests.get(pubs_query_url)
    data_request_data = data_request.text
    sc = data_request.status_code
    if sc == 200:
        global pub_id_array
        pub_id_array = []
        soup = bs4.BeautifulSoup(data_request_data, 'xml')
        pub_ids = soup.find_all('Id')
        records_data = open("# text file.txt", 'w', encoding = 'UTF-8')
        for pub_id in pub_ids:
            pub_id_text = pub_id.text
            pub_id_array += [pub_id_text]
            records_data.write(str(pub_id_text) + '\n')
        pub_id_array = sorted(set(pub_id_array))
        records_data.close()
        print(str(len(pub_id_array)) + ' PubMed IDs received')
        sleep(2)
        fetch_records_file()
    else:
        print('Access Denied: ' + str(sc) + ' error')
        quit

def fetch_records_file():
    global total_pubs_array
    total_pubs_array = []
    print('\n' + '    2. Getting records in batches of 200: ')
    for i in range(0, len(pub_id_array), 200):
        list_chunk = pub_id_array[i:i + 200]
        print('        Getting data for records ' + str(i + 1) + ' to ' + str(i + 200) + ': ', end = '')
        pub_ids_for_fetch = ','.join(list_chunk)
        record_request = requests.get(pub_records_URL_start + pub_ids_for_fetch)
        global record_request_data
        global pub_record_id_array
        sc = record_request.status_code
        if sc == 200:
            records_data = open(xml_dir + 'Records_' + str(i + 1) + '_to_' + str(i + 200) + '.xml', 'w', encoding = 'UTF-8')
            record_request_data = record_request.text
            pub_record_id_array = []
            records_data.write(record_request_data)
            soup = bs4.BeautifulSoup(record_request_data, 'xml')
            pub_id_sections = soup.find_all('MedlineCitation')
            for pub_id_section in pub_id_sections:
                pub_record_id = pub_id_section.find('PMID') # do not use 'find_all' because PMIDs for corrections appear within the same tag
                pub_record_id_text = pub_record_id.text
                pub_record_id_array += [pub_record_id_text]
                total_pubs_array += [pub_record_id_text]
            print(str(len(pub_record_id_array)) + ' PubMed records received')
            records_data.close()
        else:
            print('Access Denied: ' + str(sc) + ' error')
            quit
        sleep(1)


def parse_data(): # parse the data into arrays
    print('\n' + '    3. Parsing data from ' + str(len(xml_list)) + ' files: ')
    month_table = [['01','Jan'],['02','Feb'],['03','Mar'],['04','Apr'],['05','May'],['06','Jun'],['07','Jul'],['08','Aug'],['09','Sep'],['10','Oct'],['11','Nov'],['12','Dec']]
    global new_pmids
    new_pmids = []
    global delta_pmids
    delta_pmids = []
    global one_to_one
    global single_authors
    global single_terms
    one_to_one = []
    single_authors = []
    single_terms = []
    global file_num
    file_num = 0
    for xml_file in xml_list:
        file_num += 1
        with open(xml_file, 'r', encoding = 'UTF-8') as xml_data: # remove this line when running end-to-end
            soup = bs4.BeautifulSoup(xml_data, 'xml') # replace 'xml_data' with 'record_request_data' when running end-to-end
            pubs = soup.find_all('PubmedArticle')
            sys.stdout.write('\r')
            sys.stdout.flush()
            sys.stdout.write('        Files parsed: ' + str(file_num))
            sys.stdout.flush()
            for pub in pubs:
        # reset all lists and arrays for each PMID encountered            pmid = ''
                abstract = ''
                address = ''
                address_split = []
                addresses = ''
                au_a_reserve = ''
                au_addr_list = []
                au_address = ''
                au_orcid_list = []
                author = ''
                author_list_split = []
                author_n_address = ''
                author_n_orcid = ''
                authors = ''
                cois = ''
                doi = ''
                doi_split = ''
                email = ''
                fau_lau = ''
                full_author = ''
                full_authors = []
                grantid = ''
                grantids = []
                grants = ''
                indication = []
                indications = ''
                mesh = ''
                mesh_terms = ''
                meshes = []
                orcid = ''
                orcid_id = ''
                orig_title = ''
                other_t = ''
                other_terms = ''
                other_ts = []
                p_type = ''
                p_types = []
                pmc = ''
                pmid_link = ''
                pub_date = ''
                pub_source = ''
                pub_type = ''
                r_t_final = ''
                r_t_list = []
                r_t_source = ''
                r_t_type = ''
                reference = ''
                short_author = ''
                short_authors = []
                short_authors_list = ''
                source_split = ''
                status = ''
                title = ''
                title_with_URL = ''
                trialid = ''
                trialids = []
                trials = ''
        # start defining data for file
                pmid = pub.find('PMID').text
                pmid_link = str('https://pubmed.ncbi.nlm.nih.gov/' + pmid + '/')
                new_pmids.append(pmid)
                title = pub.find('ArticleTitle').text.replace('[', '').replace(']', '')
                orig_title_exists = pub.find('VernacularTitle')
                if orig_title_exists:
                    orig_title = orig_title_exists.text
                    title = title + ' [Original Title: ' + orig_title + ']'
                medl_cit = pub.find('MedlineCitation')
                status = medl_cit.get('Status')
                author_area = pub.find('AuthorList')
                if author_area:
                    authors_list = author_area.find_all('Author')
                    for author_name in authors_list:
                        orcid_final = ''
                        orcid = ''
                        orcid_id = ''
                        author_n_address = ''
                        author_n_orcid = ''
                        if author_name.find('LastName'):
                            author_ln = author_name.find('LastName').text
                        if author_name.find('ForeName'):
                            author_fn = author_name.find('ForeName').text
                        if author_name.find('Initials'):
                            author_initials = author_name.find('Initials').text
                        full_author = author_ln + ', ' + author_fn
                        short_author = author_ln + ' ' + author_initials
                        if author_name.find('Suffix'):
                            author_suffix = author_name.find('Suffix').text
                            full_author = author_ln + ', ' + author_fn + ' ' + author_suffix
                            short_author = author_ln + ' ' + author_initials + ' ' + author_suffix
                        author_n_address += full_author
                        author_n_orcid += full_author
                        short_authors.append(short_author)
                        full_authors.append(full_author)
                        au_affiliations = author_name.find_all('AffiliationInfo')
                        for au_affiliation in au_affiliations:
                            au_address = au_affiliation.find('Affiliation').text
                            author_n_address += ('~' + au_address)
                            au_addr_list.append(author_n_address)# only creates list if authors have an address
                        au_identifiers = author_name.find_all('Identifier')
                        for au_identifier in au_identifiers:
                            if au_identifier.get('Source') == "ORCID" and au_identifier.text[0:2] != 's0':
                                if 'orcid.org' not in au_identifier.text:
                                    orcid = 'https://orcid.org/' + str(au_identifier.text)
                                else:
                                    orcid = str(au_identifier.text.replace('http:', 'https:'))
                        author_n_orcid += ('~' + orcid)
                        au_orcid_list.append(author_n_orcid)
                    authors = '; '.join(full_authors)
                    short_authors_list = ', '.join(short_authors)
                pub_date_section = pub.find('PubDate')
                pub_year =''
                if pub_date_section.find('Year'):
                    pub_year = pub_date_section.find('Year').text
                pub_mo = ''
                if pub_date_section.find('Month'):
                    pub_month = str(pub_date_section.find('Month').text)
                    pub_mo = pub_month
                    if len(pub_month) == 2:
                        for d in range(len(month_table)):
                            if month_table[d][0] == pub_month:
                                pub_mo = month_table[d][1]
                pub_day = ''
                if pub_date_section.find('Day'):
                    pub_day = str(int(pub_date_section.find('Day').text))
                pub_date = ' '.join(filter(None,[str(pub_year), str(pub_mo), str(pub_day)]))
                if pub_date_section.find('MedlineDate'):
                    pub_date = pub_date_section.find('MedlineDate').text
                e_pub_date = ''
                e_article_mo = ''
                e_article_date_exists = pub.find('ArticleDate', {'DateType':"Electronic"})
                if e_article_date_exists:
                    e_article_year = str(e_article_date_exists.find('Year').text)
                    e_article_month = str(e_article_date_exists.find('Month').text)
                    for d in range(len(month_table)):
                        if month_table[d][0] == e_article_month:
                            e_article_mo = month_table[d][1]
                    e_article_day = str(int(e_article_date_exists.find('Day').text))
                    e_pub_date = ' '.join(filter(None,[e_article_year, e_article_mo, e_article_day]))
                article_type_tag = pub.find('Article').get('PubModel')
                article_list = pub.find('ArticleIdList')
                article_ids = article_list.find_all('ArticleId')
                for article_id in article_ids:
                    if article_id.get("IdType") == 'pmc':
                        pmc = 'https://www.ncbi.nlm.nih.gov/pmc/articles/' + article_id.text + '/'
                    if article_id.get("IdType") == 'doi':
                        doi = 'https://doi.org/' + article_id.text
                        doi_text = article_id.text
                article_sections = pub.find_all('Article')
                for article_section in article_sections:
                    if article_section.get('PubModel'):
                        journal_abbr = article_section.find('ISOAbbreviation').text
                        journal_abbr_string = journal_abbr + '. '
                        if article_section.find('Volume'):
                            journal_volume_string = ';' + str(article_section.find('Volume').text)
                        else:
                            journal_volume_string = ''
                        if article_section.find('Issue'):
                            journal_issue = str(article_section.find('Issue').text)
                            journal_issue_string = '(' + journal_issue + ')'
                        else:
                            journal_issue_string = ''
                        if article_section.find('MedlinePgn'):
                            journal_pgs_string = ':' + str(article_section.find('MedlinePgn').text)
                        elif article_section.find('ELocationID', {'EIdType':"pii"}):
                            journal_pgs_area = article_section.find('ELocationID')
                            journal_pgs_string = ':' + journal_pgs_area.text
                        else:
                            journal_pgs_string = ''
                        reference_compiled =  ''
                        website_data = ''
                        reference_compiled =  ''.join(filter(None,[journal_abbr_string, pub_date, journal_volume_string, journal_issue_string, journal_pgs_string])) + '.'
                        website_data = str(short_authors_list + '. ' + title + ' ' + reference_compiled + ' doi: ' + doi_text + '. PMID: ' + pmid)
                        if article_section.get('PubModel') == 'Electronic-eCollection':
                            reference_compiled =  ''.join(filter(None,[journal_abbr_string, e_pub_date, journal_volume_string, journal_issue_string, journal_pgs_string])) + '.'
                            website_data = str(short_authors_list + '. ' + title + ' ' + reference_compiled + ' doi: ' + doi_text + '. eCollection ' + pub_date + '. PMID: ' + pmid)
                        elif article_section.get('PubModel') == 'Print-Electronic':
                            if pub_date != e_pub_date:
                                reference_compiled =  ''.join(filter(None,[journal_abbr_string, pub_date, journal_volume_string, journal_issue_string, journal_pgs_string])) + '.'
                                website_data = str(short_authors_list + '. ' + title + ' ' + reference_compiled + ' doi: ' + doi_text + '. Epub ' + e_pub_date + '. PMID: ' + pmid)
                        # print(str(pmid) + "  |   " + str(reference_compiled) + "  |   " + str(website_data))
                if pub.find('CoiStatement'):
                    cois = pub.find('CoiStatement').text
                trials_section = pub.find('DataBankList')
                if trials_section:
                    trials_banks = trials_section.find_all('DataBank')
                    for trials_bank in trials_banks:
                        trial_bank_name = trials_bank.find('DataBankName').text
                        access_no_list = trials_bank.find('AccessionNumberList')
                        access_numbers = access_no_list.find_all('AccessionNumber')
                        for access_number in access_numbers:
                            access_id = access_number.text
                            trialid = trial_bank_name + '/' + access_id
                            trialids.append(trialid)
                    trials = '; '.join(trialids)
                grant_section = pub.find('GrantList')
                if grant_section:
                    grants = grant_section.find_all('Grant')
                    for grant in grants:
                        grant_number = ''
                        grant_acronym = ''
                        grant_agency = ''
                        grant_country = ''
                        if grant.find('GrantID'):
                            grant_number = grant.find('GrantID').text.strip()
                        if grant.find('Acronym'):
                            grant_acronym = grant.find('Acronym').text
                        if grant.find('Agency'):
                            grant_agency = grant.find('Agency').text
                        if grant.find('Country'):
                            grant_country = grant.find('Country').text
                        grantid = '/'.join(filter(None,[grant_number, grant_acronym, grant_agency, grant_country]))
                        grantids.append(grantid)
                    grants = '; '.join(grantids)
                pub_type_list = pub.find('PublicationTypeList')
                pub_types = pub_type_list.find_all('PublicationType')
                for pub_type in pub_types:
                    p_type = pub_type.text
                    p_types.append(p_type)
                pub_type = '; '.join(p_types)
                if 'Clinical Trial' in pub_type or 'Study' in pub_type or 'Case Report' in pub_type:
                    r_t_type = 'Clinical Research (Publication Type)'
                    r_t_source = 'Publication Type'
                mesh_list = pub.find('MeshHeadingList')
                if mesh_list:
                    mesh_records = mesh_list.find_all('MeshHeading')
                    for mesh_record in mesh_records:
                        mesh_descriptor_section = mesh_record.find('DescriptorName')
                        if mesh_descriptor_section.get('MajorTopicYN') == 'Y':
                            mesh_heading = '*' + mesh_descriptor_section.text
                        else:
                            mesh_heading = mesh_descriptor_section.text
                        mesh_qualifier_sections = mesh_record.find_all('QualifierName')
                        if mesh_qualifier_sections:
                            mesh_subs = []
                            for mesh_qualifier_section in mesh_qualifier_sections:
                                if mesh_qualifier_section.get('MajorTopicYN') == 'Y':
                                    mesh_subheading = '*' + mesh_qualifier_section.text
                                    mesh_subs.append(mesh_subheading)
                                else:
                                    mesh_subheading = '' + mesh_qualifier_section.text
                                    mesh_subs.append(mesh_subheading)
                            mesh_qualifiers = '/'.join(mesh_subs)
                            mesh = mesh_heading + '/' + mesh_qualifiers
                        else:
                            mesh = mesh_heading
                        meshes.append(mesh)
                    mesh_terms = '; '.join(meshes)
                    if len(indication) == 0:
                        for b in range(len(dictionary)):
                            if dictionary[b][1].lower() in mesh_terms.lower():
                                indication.append(str(dictionary[b][0]) + " ('" + str(dictionary[b][1]) + "' in MeSH Terms)")
                    if r_t_source == '':
                        for e in range(len(type_dict)):
                            if type_dict[e][0].lower() in mesh_terms.lower():
                                r_t_list.append(str(type_dict[e][1]) + " ('" + str(type_dict[e][0]) + "' in MeSH Terms)")
                other_term_list = pub.find('KeywordList')
                if other_term_list:
                    other_term_items = other_term_list.find_all('Keyword')
                    for other_term_item in other_term_items:
                        other_t = other_term_item.text
                        other_ts.append(other_t)
                    other_terms = '; '.join(other_ts)
                    if len(indication) == 0:
                        for c in range(len(dictionary)):
                            if dictionary[c][1].lower() in other_terms.lower():
                                indication.append(str(dictionary[c][0]) + " ('" + str(dictionary[c][1]) + "' in Other Terms)")
                for a in range(len(dictionary)):
                    if dictionary[a][1].lower() in title.lower():
                        indication.append(str(dictionary[a][0]) + " ('" + str(dictionary[a][1]) + "' in Title)")
                if r_t_source == '':
                    for l in range(len(type_dict)):
                        if type_dict[l][0].lower() in title.lower():
                            r_t_list.append(str(type_dict[l][1]) + " ('" + str(type_dict[l][0]) + "' in Title)")
                abstract_area = pub.find('Abstract')
                if abstract_area:
                    abstract_texts = abstract_area.find_all('AbstractText')
                    if len(abstract_texts) > 1:
                        abstract_body = ''
                        for abstract_text in abstract_texts:
                            # print(str(pmid) + ':   ' + str(abstract_text))
                            if abstract_text.get('Label'):
                                abstract_body = abstract_body + ' ' + abstract_text.get('Label').upper() + ': ' + abstract_text.text
                            else:
                                abstract_body = abstract_body + ' ' + abstract_text.text
                        abstract = abstract_body.strip()
                    else:
                        abstract_text = abstract_area.find('AbstractText')
                        abstract = abstract_text.text
                    if len(indication) == 0:
                        for d in range(len(dictionary)):
                            if dictionary[d][1].lower() in abstract.lower():
                                indication.append(str(dictionary[d][0]) + " ('" + str(dictionary[d][1]) + "' in Abstract)")
                indications = '; '.join(indication)
                if len(au_addr_list) == 0: # if there are authors but no address, create an au address list with "-" for last two field:
                    au_addr_list =[]
                    authors_split = authors.split('; ')
                    for l in range(len(authors_split)):
                        au_addr_list.append(authors_split[l] + '~-')
                if len(short_authors_list) != 0: # get first and last authors
                    if ', ' not in (short_authors_list):
                        fau_lau = short_authors_list.split(', ')[0]
                    else:
                        fau_lau = short_authors_list.split(', ')[0] + ' / ' + short_authors_list.split(', ')[-1]
                if r_t_source == '':
                    r_t_list = list(set(r_t_list))
                    r_t_type = '; '.join(r_t_list)
                if r_t_type != '':
                    r_t_final = str(r_t_type)
                corresp_author = '' # process authors, addresses, emails and corresponding author. Create single authors array.
                if len(au_addr_list) != 0:
                    deduped_addr = [] # deduplicate the addresses for each author
                    for s in au_addr_list:
                        if not any([s in r for r in au_addr_list if s != r]):
                            deduped_addr.append(s)
                    au_addr_list = deduped_addr
                    corr_authors = [] # begin creation of list with authors and addreses
                    for f in range(len(au_addr_list)):
                        author_list_split = []
                        author = ''
                        addresses = ''
                        address_split = []
                        address = ''
                        orcid_id = ''
                        email = []
                        email_text = ''
                        author_list_split = au_addr_list[f].split('~', 1) # separate author from addresses
                        author = author_list_split[0]
                        for n in range(len(au_orcid_list)):
                            au_orcid_split = au_orcid_list[n].split('~')
                            try:
                                if author == au_orcid_split[0]:
                                    orcid_id = str(au_orcid_split[1]).strip()
                            except IndexError:
                                orcid_id = ''
                        addresses = author_list_split[1]
                        try:
                            address_split = addresses.split('~') # separate individual addreses
                            for g in range(len(address_split)):
                                address = address_split[g]
                                if '@' in address:
                                    email = re.findall('\S+@\S+', address)
                                    email_text = ';'.join(email).strip('.').lower()
                                    corr_authors.append(author)
                                single_authors.append(pmid + '!~' + author + '!~' + address + '!~' + email_text + '!~' + orcid_id)
                        except IndexError:
                            address = addresses
                            if '@' in address:
                                email = re.findall('\S+@\S+', address)
                                email_text = ';'.join(email).strip('.')
                                corr_authors.append(author)
                            single_authors.append(pmid + '!~' + author + '!~' + address + '!~' + email_text + '!~' + orcid_id)
                    corresp_author = '; '.join(list(set(corr_authors))) # used in one_to_one array
                if len(mesh_terms) != 0:
                    mesh_split = []
                    mesh_split = mesh_terms.split('; ')
                    for h in range(len(mesh_split)):
                        single_terms.append(pmid + '!~' + mesh_split[h])
                one_to_one.append(pmid + '!~' + title + '!~' + authors + '!~' + corresp_author + '!~' + reference_compiled + '!~' + indications + '!~' + r_t_final + '!~' + pub_date + '!~' + doi + '!~' + pmc + '!~' + pub_type + '!~' + status + '!~' + trials + '!~' + grants + '!~' + cois + '!~' + abstract + '!~' + website_data + '!~' + pmid_link + '!~' + fau_lau)
    print('\n' + '        ' + str(len(new_pmids)) + ' publications parsed')


    write_to_xlsx()
        







def other_info(text):

    load_dotenv()

    chat = ChatOpenAI(openai_api_key=str('OpenAI API key'))

    summary = [
        SystemMessage(content='''Summarize the article in about 4 to 5 sentences.'''),
        HumanMessage(content=text[0:4000])
    ]


    drug = [
        SystemMessage(content='''What is the name of the drug that was applied in this experiment. State only the drug and nothing else. If no drug was applied, state "No Drug Was Applied".'''),
        HumanMessage(content=text[0:4000])
    ]


    target = [
        SystemMessage(content='''What is the name of the target of the drug. State only the name of the target and nothing else. If no target was used, state "No Target".'''),
        HumanMessage(content=text[0:4000])
    ]

    effect = [
        SystemMessage(content='''What was the effect of the drug on the target. If there was either no drug identified or no target identified, state "N/A".'''),
        HumanMessage(content=text[0:4000])

    ]

    title = [
        SystemMessage(content='''What is the title of the article? Return nothing BUT the title'''),
        HumanMessage(content=text[0:4000])
    ]
    
    # ML, Treatment Cycle, Medical Indication, Keywords, Title
    return chat(summary).content, chat(drug).content, chat(target).content, chat(effect).content, chat(title).content


def write_to_xlsx(): # called on at end of parse_data(), if run_delta = n
    print('\n' + '    4. Creating an Excel spreadsheet of all publications')
    workbook = xlsxwriter.Workbook(xlsx_file)
    workbook.formats[0].set_font_size(10)
    workbook.formats[0].set_font_name('Arial')
    worksheet = workbook.add_worksheet('One-to-one data')
    format_header = workbook.add_format({'font_color': 'blue', 'font_name': 'Arial', 'font_size': '10', 'valign': 'top'})
    format_wrap = workbook.add_format({'font_name': 'Arial', 'font_size': '10', 'valign': 'top', 'text_wrap': True})
    format_top = workbook.add_format({'font_name': 'Arial', 'font_size': '10', 'valign': 'top'})
    format_link = workbook.add_format({'font_color': 'blue', 'font_name': 'Arial', 'font_size': '10', 'valign': 'top', 'underline': True})
    format_tim = workbook.add_format({'font_name': 'Helvetica', 'font_size': '12', 'valign': 'top'})
    worksheet.set_row(0, None, format_header)
    worksheet.freeze_panes(1, 4)
    worksheet.set_column(0, 0, 10)
    worksheet.set_column(1, 1, 12)
    worksheet.set_column(2, 2, 12)
    worksheet.set_column(3, 3, 40)
    worksheet.set_column(4, 4, 30)
    worksheet.set_column(5, 5, 29)
    worksheet.set_column(6, 6, 20)
    worksheet.set_column(7, 7, 20)
    worksheet.set_column(8, 8, 22)
    worksheet.set_column(9, 9, 20)
    worksheet.set_column(10, 10, 17)
    worksheet.set_column(11, 11, 17)
    worksheet.set_column(12, 12, 19)
    worksheet.set_column(13, 13, 19)
    worksheet.set_column(14, 14, 15)
    worksheet.set_column(15, 15, 22)
    worksheet.set_column(16, 16, 10)
    worksheet.set_column(17, 17, 15)
    worksheet.set_column(18, 18, 40)
    worksheet.set_column(19, 19, 34)
    worksheet.set_column(20, 20, 34)
    worksheet.set_column(21, 20, 34)
    worksheet.set_column(22, 20, 34)
    worksheet.set_column(23, 20, 34)
    worksheet.set_column(24, 20, 34)
    
    row_no = 0
    worksheet.write_string(0, 0, 'PubMed ID')
    worksheet.write_string(0, 1, 'DOI Link')
    worksheet.write_string(0, 2, 'PMC Link')
    worksheet.write_string(0, 3, 'Article Title')
    worksheet.write_string(0, 4, 'Author List')
    worksheet.write_string(0, 5, 'First / Last Author')
    worksheet.write_string(0, 6, 'Corresponding Author(s)')
    worksheet.write_string(0, 7, 'Article Reference')
    worksheet.write_string(0, 8, 'Potential Indications')
    worksheet.write_string(0, 9, 'Potential R/T Types')
    worksheet.write_string(0, 10, 'Publication Date')
    worksheet.write_string(0, 11, 'Publication Type')
    worksheet.write_string(0, 12, 'Publication Status')
    worksheet.write_string(0, 13, 'Clinical Trial ID(s)')
    worksheet.write_string(0, 14, 'Grant(s)')
    worksheet.write_string(0, 15, 'Conflicts of Interest')
    worksheet.write_string(0, 16, 'Abstract')
    worksheet.write_string(0, 17, 'Website Page')
    worksheet.write_string(0, 18, 'Consolidated Web Info')
    worksheet.write_string(0, 19, 'Pubmed Link')

    # API Generated Classifications
    worksheet.write_string(0, 20, 'Summary')
    worksheet.write_string(0, 21, 'Drug Applied')
    worksheet.write_string(0, 22, 'Targer')
    worksheet.write_string(0, 23, 'Effect on Target')

    row_no = 1
    for i in range(len(one_to_one)):
        data_line = one_to_one[i].split('!~')
        pmc_text = ''
        if data_line[9]:
            pmc_split = data_line[9].split('/')
            pmc_text = pmc_split[-2]
        col = 0
        worksheet.write_url(row_no, col, data_line[17], format_top, string=data_line[0]) # 0 PMID with link
        worksheet.write_url(row_no, col + 1, data_line[8], format_link) # 1 DOI
        if data_line[9]:
            worksheet.write_url(row_no, col + 2, data_line[9], format_link, pmc_text) # 2 PMC Full Text Link
        else:
            worksheet.write_string(row_no, col + 2, '', format_top) # 2 blank if PMC Full Text Link is null
        worksheet.write_string(row_no, col + 3, data_line[1], format_wrap) # 3 Title
        worksheet.write_string(row_no, col + 4, data_line[2], format_wrap) # 4 Author list
        worksheet.write_string(row_no, col + 5, data_line[18], format_wrap) # 5 FAU / LAU
        worksheet.write_string(row_no, col + 6, data_line[3], format_wrap) # 6 Correponding Author
        worksheet.write_string(row_no, col + 7, data_line[4], format_wrap) # 7 Reference
        worksheet.write_string(row_no, col + 8, data_line[5], format_top) # 8 Potential Indications
        worksheet.write_string(row_no, col + 9, data_line[6], format_top) # 9 Potential R/T Type
        worksheet.write_string(row_no, col + 10, data_line[7], format_top) # 10 Publication Date
        worksheet.write_string(row_no, col + 11, data_line[10], format_top) # 11 Publication Type
        worksheet.write_string(row_no, col + 12, data_line[11], format_top) # 12 Pub Status
        worksheet.write_string(row_no, col + 13, data_line[12], format_top) # 13 Clinical Trial IDs
        worksheet.write_string(row_no, col + 14, data_line[13], format_top) # 14 Grants
        worksheet.write_string(row_no, col + 15, data_line[14], format_top) # 15 Conflict of Interest
        worksheet.write_string(row_no, col + 16, data_line[15], format_top) # 16 Abstract
        worksheet.write_string(row_no, col + 17, '', format_top) # 17 Tim website Page
        worksheet.write_string(row_no, col + 18, data_line[16], format_tim) # 18 Tim consolidated info for website
        worksheet.write_string(row_no, col + 19, data_line[17], format_top) # 19 Tim PMID link
       
        worksheet.write_string(row_no, col + 21, other_info(data_line[15])[0], format_wrap) # 21 Summary
        worksheet.write_string(row_no, col + 22, other_info(data_line[15])[1], format_wrap) # 22 Drug
        worksheet.write_string(row_no, col + 23, other_info(data_line[15])[2], format_wrap) # 23 Target
        worksheet.write_string(row_no, col + 24, other_info(data_line[15])[3], format_wrap) # 24 Effect on target
        row_no += 1

    worksheet = workbook.add_worksheet('Authors')
    worksheet.set_row(0, None, format_header)
    worksheet.freeze_panes(1, 0)
    worksheet.set_column(0, 0, 10)
    worksheet.set_column(1, 1, 30)
    worksheet.set_column(2, 2, 80)
    worksheet.set_column(3, 3, 40)
    worksheet.set_column(4, 4, 31)
    row_no = 0
    worksheet.write_string(0, 0, 'PubMed ID')
    worksheet.write_string(0, 1, 'Author Name')
    worksheet.write_string(0, 2, 'Author Affiliation')
    worksheet.write_string(0, 3, 'Author Email')
    worksheet.write_string(0, 4, 'ORCID_ID URL')
    row_no = 1
    for j in range(len(single_authors)):
        data_line = single_authors[j].split('!~')
        col = 0
        worksheet.write_string(row_no, col, data_line[0], format_top)
        worksheet.write_string(row_no, col + 1, data_line[1], format_top)
        worksheet.write_string(row_no, col + 2, data_line[2], format_wrap)
        worksheet.write_string(row_no, col + 3, data_line[3], format_top)
        if data_line[4]:
            worksheet.write_url(row_no, col + 4, data_line[4], format_link)
        else:
            worksheet.write_string(row_no, col + 4, data_line[4], format_top)
        row_no += 1

    worksheet = workbook.add_worksheet('MeSH Terms')
    worksheet.set_row(0, None, format_header)
    worksheet.freeze_panes(1, 0)
    worksheet.set_column(0, 0, 10)
    worksheet.set_column(1, 1, 20)
    row_no = 0
    worksheet.write_string(0, 0, 'PubMed ID')
    worksheet.write_string(0, 1, 'MeSH Term')
    row_no = 1
    for k in range(len(single_terms)):
        data_line = single_terms[k].split('!~')
        col = 0
        worksheet.write_string(row_no, col, data_line[0])
        worksheet.write_string(row_no, col + 1, data_line[1])
        row_no += 1
    workbook.close()

    
fetch_pub_ids()
xml_list = glob.glob(xml_files)
parse_data()
print(xml_list)
end_time = datetime.now()
print('')
print('\n' + 'The program completed in ' + str(end_time.replace(microsecond=0)-start_time.replace(microsecond=0)) + ', including ' + str(file_num + 4) + ' seconds of sleep time. Here is the summary: ')
print('    Records identified:           ' + str(len(pub_id_array)))
print('    Records received from PubMed: ' + str(len(total_pubs_array)))
print('    Records parsed:               ' + str(len(new_pmids)))

print(f'XML FILES {xml_list[0]}')
